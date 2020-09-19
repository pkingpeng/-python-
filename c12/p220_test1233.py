import openpyxl

wb = openpyxl.load_workbook('../resource/excel/example.xlsx')
sheet = wb['Sheet1']

print(sheet['A1'])
print(sheet['A1'].value)

temp = sheet['B1']
print(temp.value)

print('Row %s, Column %s is %s.' % (temp.row, temp.column, temp.value))
print('Cell %s is %s.' % (temp.coordinate, temp.value))

print(sheet['C1'].value)

print('-' * 50)

m = sheet.cell(row=1, column=2)
print(m)
print(m.value)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print('-' * 50)

print(sheet.max_row)
print(sheet.max_column)
print(sheet.calculate_dimension())

"""
<Cell 'Sheet1'.A1>
2015-04-05 13:34:02
Apples
Row 1, Column 2 is Apples.
Cell B1 is Apples.
73
--------------------------------------------------
<Cell 'Sheet1'.B1>
Apples
1 Apples
3 Pears
5 Apples
7 Strawberries
--------------------------------------------------
7
3
A1:C7
"""

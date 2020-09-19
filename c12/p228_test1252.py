import openpyxl

wb = openpyxl.Workbook()
print(wb.sheetnames)

wb.create_sheet()
print(wb.sheetnames)

wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)

wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)

wb.remove(wb['Middle Sheet'])
wb.remove(wb['Sheet1'])
print(wb.sheetnames)

"""
['Sheet']
['Sheet', 'Sheet1']
['First Sheet', 'Sheet', 'Sheet1']
['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']
['First Sheet', 'Sheet']
"""

"""
https://openpyxl.readthedocs.io/en/stable/formula.html
参考了半天没有找到解析出公式值的办法，直接打开 Excel 可以看到数字是对的。
"""
import openpyxl

xlsx = '../resource/excel/writeFormula.xlsx'

wb = openpyxl.Workbook()
sheet = wb['Sheet']

sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'

wb.save(xlsx)

print('Generate Success')

wbFormulas = openpyxl.load_workbook(xlsx)
sheet = wbFormulas.active
print(sheet['A3'].value)


wbDataOnly = openpyxl.load_workbook(xlsx, data_only=True)
sheet1 = wbDataOnly.active
print(sheet1['A3'].value)

"""
Generate Success
=SUM(A1:A2)
None
"""

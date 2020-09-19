#! python3
# Corrects coats in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('../resource/excel/produceSales.xlsx')
sheet = wb['Sheet']

price_update_dic = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

print('Start update produce...')
count = 0

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in price_update_dic:
        sheet.cell(row=rowNum, column=2).value = price_update_dic[produceName]
        count += 1

wb.save('../resource/excel/output/updateProduceSales.xlsx')
print('End update produce and update %d record.' % count)

"""
Start update produce...
End update produce and update 1823 record.
"""

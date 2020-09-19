"""
"""
import openpyxl

xlsx = '../resource/excel/merge.xlsx'

wb = openpyxl.Workbook()
sheet = wb['Sheet']

sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells'

wb.save(xlsx)

print('Generate Success')

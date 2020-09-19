"""
https://stackoverflow.com/questions/36721232/importerror-cannot-import-name-get-column-letter
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

wb = openpyxl.load_workbook('../resource/excel/example.xlsx')
sheet = wb['Sheet1']

print(sheet.max_column)
print(get_column_letter(sheet.max_column))

print(column_index_from_string('A'))
print(column_index_from_string('AA'))

"""
A
B
AA
AHP
3
C
1
27
"""

import openpyxl

wb = openpyxl.Workbook()
print(wb.sheetnames)

sheet = wb.active
print(sheet.title)

sheet.title = 'Spam Bacocn Eggs Sheet'
print(wb.sheetnames)

print('-' * 50)


wb = openpyxl.load_workbook('../resource/excel/example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save('../resource/excel/output/example_copy.xlsx')
print('sava success...')

"""
['Sheet']
Sheet
['Spam Bacocn Eggs Sheet']
--------------------------------------------------
sava success...
"""

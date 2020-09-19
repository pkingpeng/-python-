#!
#  /Users/moqi/Downloads/excelSpreadsheets

import csv
import openpyxl
import os

os.chdir('/Users/moqi/Downloads/excelSpreadsheets')
os.makedirs('excel2csv', exist_ok=True)

for fileName in os.listdir('.'):
    if not fileName.endswith('.xlsx'):
        continue

    wb = openpyxl.load_workbook(fileName, data_only=True)

    for sheetName in wb.sheetnames:
        sheet = wb[sheetName]

        # excelName 去掉 .xlsx 后缀
        csvName = fileName[:-5] + '-' + sheetName + '.csv'
        outCsvFileObj = open(os.path.join('excel2csv', csvName), 'w')
        writer = csv.writer(outCsvFileObj)

        for rowNum in range(1, sheet.max_row + 1):
            rowData = []

            for colNum in range(1, sheet.max_column + 1):
                rowData.append(sheet.cell(row=rowNum, column=colNum).value)

            writer.writerow(rowData)

        outCsvFileObj.close()

    wb.close()

print('Done.')

"""
遇到的问题: https://blog.csdn.net/WinterShiver/article/details/103443300
1. 报错：zipfile.BadZipFile: File is not a zip file
   可能是 Excel 未正常关闭导致的，先关闭相关文件夹下所有 Excel 实例。解决
"""
